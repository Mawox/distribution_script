from openpyxl import Workbook, load_workbook 
from pandas import DataFrame
#from openpyxl.worksheet import Cell
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.worksheet.page import PrintOptions



def load_xlsx(file_path: str) -> DataFrame:

    wb = load_workbook(filename=file_path)
    ws = wb["Form Responses 1"]
    df = DataFrame(ws.values)
    df.columns = df.iloc[0]

    df["skola"] = df["Z jaké jste školy?"].astype(str).replace("None", "").sum(axis=1)

    return df


def add_ws(line: int, df: DataFrame, wb: Workbook):
    df = df.iloc[line]
    ws = wb.create_sheet()

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 27

    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1

    wsprops = ws.sheet_properties
    wsprops.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
    ws.print_options = PrintOptions(horizontalCentered=True, gridLines=True)

    skola = df.loc["skola"].split("(")[0]
    cell = ws["A1"]
    cell.value = skola
    cell.font = Font(size=36)
    ws.append([""])
    osoba = df.pop(
        "Vaše jméno a příjmení (kontaktní osoba pro účely této objednávky)"
    )
    telefon = df.pop(
        "Vaše telefonní číslo (kontaktní osoba pro účely této objednávky)"
    )
    email = df.pop("Váš e-mail (kontaktní osoba pro účely této objednávky)")
    poznamka = df.pop("Jakékoliv další poznámky k objednávce či dopravě")

    cislo = df.pop("Číslo popisné")
    ulice = df.pop("Ulice")

    psc = df.pop("PSČ")
    obec = df.pop(
        "Obec (název obce nebo části obce případně městská část nebo městský obvod)"
    )

    ws.append([ulice, cislo])
    ws.append([psc, obec])

    ws.append(["Kontaktní osoba: ", osoba])
    ws.append(["Tel.: ", telefon])
    ws.append(["E-mail: ", email])
    ws.append([""])

    cell = ws["A8"]
    cell.value = poznamka
    cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells('A8:C8')
    rd = ws.row_dimensions[8]
    rd.height = 30
    ws.append([""])

    ws.append(["Název", "Autor", "ks"])
    celkem = 0
    for n in range(len(df)):
        try:
            ks = int(df[n])
        except ValueError:
            continue
        except TypeError:
            continue
        celkem += ks
        name = df.axes[0][n]

        if "; " in name:
            name = name.split("; ")[1]
            author, book = name.split(": ", 1)
        else:
            book = name
            author = ""
        ws.append([book, author, ks])
    ws.append([""])
    ws.append(["", "CELKEM KS", celkem])

def create_output(df: DataFrame):
    wb = Workbook()

    for i in range(1, len(df)):
        add_ws(i, df, wb)
    #print(wb.active)
    wb.save(r"./data/write_only_file.xlsx")


if __name__ == "__main__":
    data_df = load_xlsx(r"./data/input.xlsx")
    create_output(data_df)

    print("END")
